#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
import pandas as pd


# In[3]:


wb = openpyxl.load_workbook('reporte.xlsx')
hoja = wb.active
filas = hoja.max_row


# In[4]:


#Lista con los nombres de los radios censales
radios = []
filasRadios = []
for i in range(1,filas):
    try:
        dato = hoja.cell(row = i, column = 1).value
        if 'AREA #' in dato:
            radio = dato.split()[-1] #Agarra solamente el número después del segundo espacio
            radios.append(radio)
            filasRadios.append(i)
        else:
            pass
    except:
            pass    


# In[5]:


#Busca las dos primera fila con radio:
for i in range(1,filas):
    try:
        dato = hoja.cell(row = i, column = 1).value
        if radios[0] in dato:
            primero = i
        elif radios[1] in dato:
            segundo = i
            break
    except:
            pass
#Cuántas filas tiene que ignorar entre uno y otro:
salto = segundo - primero


# In[6]:


#Columas del DataFrame:
columnasDF = ['link'] #'link' primero para el index. Después las categorías.
for i in range(primero+3,segundo-2):
    columnasDF.append(hoja.cell(row = i, column = 1).value)


# In[7]:


#Armo diccionario del DF
dictDF = {}

#Defino la columna de links:
dictDF['link'] = radios

#Lleno con ceros
for col in columnasDF:
    if col is not 'link':
        dictDF[col] = ['0' for i in range(len(radios))]
    
#Cambio los valoes de las otras listas del diccionario.
posicion = 0
for f in range(len(filasRadios)):
    fila = filasRadios[f]
    for i in range(fila, fila + len(columnasDF)+3):
        col = hoja.cell(row = i, column = 1).value
        if col in columnasDF:
            dictDF[col][posicion] = hoja.cell(row = i, column = 3).value
            
    posicion += 1
    #print(posicion)


# In[8]:


#Defino el DataFrame
datos = pd.DataFrame(dictDF)
datos.set_index('link', inplace = True)

#Si una columna tiene NaN, pandas la reconoce como object.
#Reemplazo por 0 en donde no hay dato
datos.fillna('0')

#Convierto las object a float64
for col in datos.columns:
    if datos[col].dtype == 'object':
        datos[col] = datos[col].astype('float64')


# In[9]:


#Guardo como csv
datos.to_csv('reporte.csv')

#Guardo como pickle
datos.to_pickle('reporte.pickle')


# In[10]:


#Guardo como excel
datos.to_excel('reporte_table.xlsx')

